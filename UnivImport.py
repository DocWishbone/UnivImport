import sys
import subprocess
from datetime import datetime

# ============================================
# Abh√§ngigkeiten pr√ºfen / bei Bedarf installieren
# ============================================

REQUIRED_MODULES = [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("numpy", "numpy"),
    ("colorama", "colorama"),
]

def ensure_dependencies():
    print("Pr√ºfe ben√∂tigte Python-Module...")
    for module_name, package_name in REQUIRED_MODULES:
        try:
            __import__(module_name)
        except ImportError:
            print(f"  Modul '{module_name}' fehlt ‚Äì installiere '{package_name}'...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
                print(f"  ‚Üí '{package_name}' wurde installiert.")
            except Exception as e:
                print(f"  Konnte '{package_name}' nicht installieren: {e}")
                print("  Bitte manuell installieren und Script erneut starten.")
                sys.exit(1)

ensure_dependencies()

# ============================================
# Imports nach erfolgreichem Dependency-Check
# ============================================

import os
import shutil
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from colorama import init, Fore, Style

init(autoreset=True)

VERSION = "2.2.6"

#########################################
# UnivImport.py - 17.12.2025 - m.ludwig #
#########################################

# Zentrale Master-Version des Scripts auf dem Netzlaufwerk
UPDATE_SOURCE = r"\\hrl.local\fs\data\MindenBruhns\MBImport\UnivImport\UnivImport.py"

# === Einstellungen ===
INPUT_MB = "Mappe1.xlsx"
INPUT_NG = "NG.xlsx"
INPUT_NEF = "NEF.csv"
OUTPUT_CSV = r"\\hrl.local\fs\data\niederegger\csv\mb\import.csv"
OUTPUT_XLSX = "import.xlsx"
START_BANNER = r"""
              *
             ***
            *****
           *******
          *********
         ***********
             |||
             |||

      UnivImport ‚Äì Ho Ho Ho, Frohe Festtage und guten Rutsch .... üéÖ
"""

# Spalten, die wir aus der Originaldatei behalten m√∂chten
columns_to_keep = [
    "Artikelnummer",
    "Benennung",
    "LHM-Nr.",
    "Charge",
    "Menge",
    "Einheit",
    "Stelltyp",
    "MHD",
    "Gesamtgewicht",
]

# Ziel-Spaltennamen f√ºr den Import
rename_map = {
    "Artikelnummer": "Artikel-Nr.",
    "Benennung": "Artikelbezeichnung",
    "LHM-Nr.": "LG ID",
    "Charge": "Charge",
    "Menge": "Menge PS",
    "Einheit": "Einheit",
    "Stelltyp": "Lademittel",
    "MHD": "MHD",
    "Gesamtgewicht": "Gewicht kg",
}

# Mapping-Tabelle f√ºr Lademittel
LADEMITTEL_MAPPING = {
    "euro": "Euro",
    "h1": "H1",
    "industrie": "Industrie",
    # Erweiterbar
}

def map_lademittel(value: str) -> str:
    """Normalisiert 'Lademittel' anhand der Mapping-Tabelle."""
    if not isinstance(value, str):
        return ""
    v = value.lower()
    for key, target in LADEMITTEL_MAPPING.items():
        if key in v:
            return target
    return ""

def ensure_latest_version():
    """
    Pr√ºft, ob auf dem Netzlaufwerk eine neuere import.py liegt.
    Wenn ja, wird sie √ºber die lokale Datei kopiert und das Script neu gestartet.
    """
    try:
        local_path = Path(__file__).resolve()
    except NameError:
        print(Fore.RED + "[Update] Konnte __file__ nicht bestimmen ‚Äì Update-Check wird √ºbersprungen.")
        return

    network_path = Path(UPDATE_SOURCE)

    # print(Fore.CYAN + f"[Update] Lokales Script : {local_path}")
    # print(Fore.CYAN + f"[Update] Master-Script : {network_path}")

    if not network_path.exists():
        print(Fore.YELLOW + "[Update] Master-Script nicht gefunden ‚Äì kein Update m√∂glich.")
        return

    try:
        net_mtime = network_path.stat().st_mtime
        local_mtime = local_path.stat().st_mtime
    except OSError as e:
        print(Fore.RED + f"[Update] Fehler beim Lesen der Dateizeitstempel: {e}")
        return

    print(Fore.CYAN + f"[Update] Lokales Datum : {datetime.fromtimestamp(local_mtime)}")
    print(Fore.CYAN + f"[Update] Master-Datum : {datetime.fromtimestamp(net_mtime)}")

    # Nur wenn die Netzlaufwerk-Version wirklich neuer ist
    if net_mtime > local_mtime:
        try:
            shutil.copy2(network_path, local_path)
            print(Fore.GREEN + "\n=== UPDATE DURCHGEF√úHRT ===")
            print(Fore.GREEN + f"Neue Version von: {network_path}")
            print(Fore.GREEN + f"Aktualisiert nach: {local_path}")
            print(Fore.GREEN + "Script wird mit der neuen Version neu gestartet...")

            python_exe = sys.executable or "python"
            script_path = str(local_path)

            # Debug-Ausgabe
            print(Fore.CYAN + "[Update] Neustart-Befehl:")
            print(Fore.CYAN + f"         {python_exe!r} {script_path!r} {sys.argv[1:]}")

            # Neuen Prozess mit sauber getrennten Argumenten starten
            subprocess.Popen([python_exe, script_path] + sys.argv[1:])

            print(Fore.CYAN + "[Update] Alter Prozess wird beendet.")
            sys.exit(0)

        except Exception as e:
            print(Fore.RED + f"[Update] Fehler beim Aktualisieren der Script-Version: {e}")
            print(Fore.YELLOW + "Es wird mit der aktuellen lokalen Version fortgefahren.")
            return
    else:
        print(Fore.GREEN + "[Update] Lokale Version ist aktuell (oder neuer). Kein Update n√∂tig.")

import pandas as pd

def _find_col(columns, needle: str) -> str:
    needle = needle.lower()
    for c in columns:
        if c is None:
            continue
        if needle in str(c).strip().lower():
            return c
    raise KeyError(f"Keine Spalte gefunden, die '{needle}' enth√§lt. Vorhanden: {list(columns)}")

import re

def _norm_match(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()

    # erste Zahl finden: 10030004.112 (Punkt optional)
    m = re.search(r"\d+(?:\.\d+)?", s)
    return m.group(0) if m else ""

def load_artikelmap_from_excel_fuzzy(path: str, kunde_filter: str | None = None) -> dict:
    xls = pd.ExcelFile(path)

    last_err = None
    for sheet in xls.sheet_names:
        try:
            art = pd.read_excel(path, sheet_name=sheet)

            col_match = _find_col(art.columns, "match")
            col_bez   = _find_col(art.columns, "bezeichnung")
            col_kunde = _find_col(art.columns, "kunde")

            if kunde_filter:
                mask = art[col_kunde].astype(str).str.lower().str.contains(str(kunde_filter).lower(), na=False)
                art = art[mask].copy()

            art = art.dropna(subset=[col_match, col_bez]).copy()

            keys = art[col_match].apply(_norm_match)
            vals = art[col_bez].astype(str).str.strip()

            artikel_map = {}
            for k, v in zip(keys, vals):
                if not k:
                    continue
                artikel_map.setdefault(k, v)

            return artikel_map

        except Exception as e:
            last_err = e
            continue

    raise ValueError(f"Konnte in keiner Tabelle passende Spalten finden. Letzter Fehler: {last_err}")

import csv

def read_csv_robust(path: str | Path) -> tuple[pd.DataFrame, str, str]:
    # delimiter + encoding grob erkennen
    with open(path, "rb") as f:
        sample = f.read(4096)

    encoding = None
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = sample.decode(enc, errors="strict")
            encoding = enc
            break
        except Exception:
            pass
    if encoding is None:
        encoding = "latin1"
        text = sample.decode(encoding, errors="ignore")

    try:
        dialect = csv.Sniffer().sniff(text, delimiters=";,|\t,")
        sep = dialect.delimiter
    except Exception:
        sep = ";"

    df = pd.read_csv(path, sep=sep, encoding=encoding, dtype=str)
    return df, sep, encoding

def main():
    # =====================================
    # Kundenauswahl
    # =====================================
    KUNDEN = {
        "0": ("BEENDEN", None),
        "1": ("MB", "standard"),
        "2": ("NG", "ng"),
        "3": ("NEF","nef"),
    }

    print("\nKunde ausw√§hlen:")
    for key, (label, _) in KUNDEN.items():
        print(f"  {key} = {label}")

    while True:
        choice = input("Auswahl (0/1/2/3): ").strip()
        if choice in KUNDEN:
            break
        print(Fore.RED + "Ung√ºltige Auswahl ‚Äì bitte 0, 1 oder 2 eingeben.")

    label, kunde = KUNDEN[choice]
    if kunde is None:
        print(Fore.YELLOW + "Vorgang abgebrochen.")
        return   # main() sauber beenden

    print(f"‚Üí Gew√§hlt: {label}\n")
    
    input_file = INPUT_NEF if kunde == "nef" else (INPUT_NG if kunde == "ng" else INPUT_MB)
    if not Path(input_file).exists():
        raise FileNotFoundError(f"Eingabedatei nicht gefunden: {input_file}")

    # =====================================
    # Einlesen (abh√§ngig vom Kunden)
    # =====================================
    extra_excel_line = ""

    if kunde == "ng":
        # --- NG: positionsbasiert einlesen ---
        df_raw = pd.read_excel(input_file, header=None)
        df_raw = df_raw.dropna(subset=[0]).copy()

        df = pd.DataFrame({
            "Artikelnummer": df_raw[0].astype(str).str.strip(),
            "Benennung": "",
            "LHM-Nr.": df_raw[7],
            "Charge": df_raw[6],
            "Menge": 1,
            "Einheit": "BigBag",
            "Stelltyp": "",
            "MHD": df_raw[5],
            "Gesamtgewicht": df_raw[1],
        })

        # Artikelstamm ‚Üí Bezeichnung
        artikel_map = load_artikelmap_from_excel_fuzzy("artikel.xlsx", kunde_filter="NG")
        df["Artikelnummer"] = df["Artikelnummer"].apply(_norm_match)
        df["Benennung"] = df["Artikelnummer"].map(artikel_map).fillna("")

    elif kunde == "nef":
        # --- NEF: CSV einlesen (Encoding/Trenner robust) ---
        df, sep, enc = read_csv_robust(input_file)
        print(Fore.CYAN + f"[NEF] CSV gelesen (sep={sep!r}, encoding={enc!r})")

        # Spaltennamen s√§ubern
        df.columns = [str(c).strip() for c in df.columns]
        
        # NEF-Spalten auf unsere Standardnamen bringen
        rename_nef = {
            "Menge Kart.": "Menge PS",
            "Bruttogewicht kg": "Gewicht kg",
        }
        df = df.rename(columns=rename_nef)

        # Pflichtspalten (so wie NEF sie wirklich liefert)
        required = ["LG ID", "Artikel-Nr.", "Artikelbezeichnung", "Menge PS", "MHD", "Charge"]
        missing_cols = [c for c in required if c not in df.columns]
        if missing_cols:
            raise ValueError(f"[NEF] Fehlende Spalten in CSV: {missing_cols}. Vorhanden: {list(df.columns)}")

        # Defaults, weil NEF diese Spalten nicht liefert
        if "Einheit" not in df.columns:
            df["Einheit"] = "UMK"
        if "Lademittel" not in df.columns:
            df["Lademittel"] = ""
        if "Gewicht kg" not in df.columns:
            df["Gewicht kg"] = ""
        if "Lagerort" not in df.columns:
            df["Lagerort"] = ""
        if "Sonstiger Text" not in df.columns:
            df["Sonstiger Text"] = ""
        
        if "Lademittel" in df.columns:
            df = df.drop(columns=["Lademittel"])
        if "Sonstiger Text" in df.columns:
            df = df.drop(columns=["Sonstiger Text"])

        # Typen wie bei euch
        df["LG ID"] = pd.to_numeric(df["LG ID"], errors="coerce").fillna(0).astype(int)
        df["Menge PS"] = pd.to_numeric(df["Menge PS"], errors="coerce").fillna(0).astype(int)

        # Gewicht putzen (falls sp√§ter gebraucht)
        gewicht_raw = df["Gewicht kg"].astype(str).str.strip()
        gewicht_raw = gewicht_raw.str.replace(r"[^0-9,\.]", "", regex=True)
        gewicht_raw = gewicht_raw.str.replace(r"\.(?=[0-9]{3}(?:$|,))", "", regex=True)
        gewicht_raw = gewicht_raw.str.replace(",", ".", regex=False)
        df["Gewicht kg"] = pd.to_numeric(gewicht_raw, errors="coerce").fillna(0).round(2)

    else:
        # --- STANDARD: Header automatisch finden ---
        df_raw = pd.read_excel(input_file, header=None)

        expected_headers = [
            "Artikelnummer",
            "Benennung",
            "LHM-Nr.",
            "Charge",
            "Menge",
            "Einheit",
            "Stelltyp",
            "MHD",
            "Gesamtgewicht",
        ]

        def norm(x):
            return str(x).strip() if pd.notna(x) else ""

        header_row_idx = None
        for idx, row in df_raw.iterrows():
            normalized_row = [norm(v) for v in row.tolist()]
            if all(h in normalized_row for h in expected_headers):
                header_row_idx = idx
                break

        if header_row_idx is None:
            raise ValueError("Konnte keine Kopfzeile mit den erwarteten Spalten√ºberschriften finden.")

        header_row = [norm(v) for v in df_raw.iloc[header_row_idx].tolist()]
        df = df_raw.iloc[header_row_idx + 1:].copy()
        df.columns = header_row


    # =====================================
    # MB/NG Aufbereitung (NEF √ºberspringen!)
    # =====================================
    if kunde != "nef":
        df = df[columns_to_keep]
        df = df.dropna(how="all")

        art_raw = df["Artikelnummer"].astype(str).str.strip()
        pattern = r"\d+(?:\.\d+)?" if kunde == "ng" else r"\d+"
        mask_art = art_raw.str.fullmatch(pattern)

        df = df[mask_art].copy()
        df["Artikelnummer"] = art_raw[mask_art]

        # MHD robust parsen
        mhd_raw = df["MHD"]
        mhd_parsed = pd.to_datetime(mhd_raw, errors="coerce", dayfirst=True)

        mask_num = mhd_raw.notna() & mhd_raw.apply(lambda x: isinstance(x, (int, float)))
        if mask_num.any():
            mhd_parsed.loc[mask_num] = pd.to_datetime(mhd_raw[mask_num].astype(float), unit="D", origin="1899-12-30")

        mask_digit_str = mhd_raw.notna() & ~mask_num & mhd_raw.astype(str).str.fullmatch(r"\d+")
        if mask_digit_str.any():
            mhd_parsed.loc[mask_digit_str] = pd.to_datetime(mhd_raw[mask_digit_str].astype(float), unit="D", origin="1899-12-30")

        df["MHD"] = mhd_parsed.dt.strftime("%d.%m.%Y")

        df = df.rename(columns=rename_map)

        df["Einheit"] = df["Einheit"].apply(lambda x: "UMK" if isinstance(x, str) and x.strip().lower() == "container" else x)
        df["Lademittel"] = df["Lademittel"].apply(map_lademittel)

        df["Menge PS"] = pd.to_numeric(df["Menge PS"], errors="coerce").fillna(0)

        gewicht_raw = df["Gewicht kg"].astype(str).str.strip()
        gewicht_raw = gewicht_raw.str.replace(r"[^0-9,\.]", "", regex=True)
        gewicht_raw = gewicht_raw.str.replace(r"\.(?=[0-9]{3}(?:$|,))", "", regex=True)
        gewicht_raw = gewicht_raw.str.replace(",", ".", regex=False)
        df["Gewicht kg"] = pd.to_numeric(gewicht_raw, errors="coerce").fillna(0)

        df = df.groupby("LG ID", as_index=False).agg({
            "Artikel-Nr.": "first",
            "Artikelbezeichnung": "first",
            "Charge": "first",
            "Menge PS": "sum",
            "Einheit": "first",
            "Lademittel": "first",
            "MHD": "first",
            "Gewicht kg": "sum",
        })

        df["Artikel-Nr."] = df["Artikel-Nr."].astype(str).str.strip()
        if kunde == "standard":
            df["Artikel-Nr."] = df["Artikel-Nr."].str.zfill(6)

        df["LG ID"] = pd.to_numeric(df["LG ID"], errors="coerce").fillna(0).astype(int)
        df["Menge PS"] = df["Menge PS"].fillna(0).astype(int)

        charge_num = pd.to_numeric(df["Charge"], errors="coerce").fillna(0).astype(int)
        df["Charge"] = charge_num.astype(str)

        df["Gewicht kg"] = df["Gewicht kg"].round(2)

    # Zusatzspalten
    df["Lagerort"] = ""
    df["Sonstiger Text"] = ""

     # === Interaktive Abfrage f√ºr Lagerort & Sonstiger Text (j/n/a/h) ===
    try:
        main_help = f"""
    {Fore.CYAN}HILFE ‚Äì Lagerort / Sonstiger Text{Style.RESET_ALL}

      {Fore.GREEN}j{Style.RESET_ALL} = Eingaben gelten f√ºr {Fore.YELLOW}alle{Style.RESET_ALL} Zeilen
      {Fore.RED}n{Style.RESET_ALL} = keine √Ñnderungen
      {Fore.CYAN}a{Style.RESET_ALL} = Auswahl einzelner Zeilen / Bereiche
      {Fore.CYAN}h{Style.RESET_ALL} = diese Hilfe anzeigen

    Auswahl ({Fore.CYAN}a{Style.RESET_ALL}):
      - Einzelne Zeilen:  {Fore.YELLOW}3, 7, 12{Style.RESET_ALL}
      - Bereiche:         {Fore.YELLOW}1-10{Style.RESET_ALL}
      - Kombinationen:    {Fore.YELLOW}1-5, 8, 12-15{Style.RESET_ALL}
      - Leere Eingabe beendet die Auswahl
    """

        select_help = f"""
    {Fore.CYAN}HILFE ‚Äì Zeilenauswahl{Style.RESET_ALL}

    Beispiele:
      1-10        ‚Üí Zeile 1 bis 10
      5,7,9       ‚Üí einzelne Zeilen
      1-5, 12-15  ‚Üí Kombinationen

    Tipps:
      - Nummern sind {Fore.YELLOW}1-basiert{Style.RESET_ALL} (so wie angezeigt)
      - Leere Eingabe beendet die Auswahl
    """

        def parse_ranges(text):
            result = []
            parts = [p.strip() for p in text.split(",") if p.strip()]
            for part in parts:
                if "-" in part:
                    start, end = part.split("-")
                    result.extend(range(int(start), int(end) + 1))
                else:
                    result.append(int(part))
            return sorted(set(result))

        # ============================
        # NEF: nur Lagerort abfragen
        # ============================
        if kunde == "nef":
            lagerort = input(
                f"{Fore.YELLOW}Lagerort (NEF) ‚Äì leer = keine √Ñnderung:{Style.RESET_ALL} "
            ).strip()
            if lagerort:
                df["Lagerort"] = lagerort

        # ============================
        # MB/NG: alter Dialog
        # ============================
        else:
            while True:
                prompt = (
                    f"M√∂chten Sie '{Fore.YELLOW}Lagerort{Style.RESET_ALL}' und "
                    f"'{Fore.YELLOW}Sonstiger Text{Style.RESET_ALL}' setzen? "
                    f"({Fore.GREEN}j{Style.RESET_ALL}=alle / "
                    f"{Fore.RED}n{Style.RESET_ALL}=keine / "
                    f"{Fore.CYAN}a{Style.RESET_ALL}=Auswahl / "
                    f"{Fore.CYAN}h{Style.RESET_ALL}=Help): "
                )
                answer = input(prompt).strip().lower()

                if answer == "h":
                    print(main_help)
                    continue
                if answer in ("j", "n", "a"):
                    break
                print(f"{Fore.RED}Ung√ºltige Eingabe ‚Äì 'h' f√ºr Hilfe.{Style.RESET_ALL}")

            if answer == "j":
                lagerort = input(f"{Fore.YELLOW}Lagerort (leer = keine √Ñnderung):{Style.RESET_ALL} ").strip()
                sonstiger_text = input(f"{Fore.YELLOW}'Sonstiger Text' (leer = keine √Ñnderung):{Style.RESET_ALL} ").strip()
                if lagerort:
                    df["Lagerort"] = lagerort
                if sonstiger_text:
                    df["Sonstiger Text"] = sonstiger_text

            elif answer == "a":
                print("\nVerf√ºgbare Zeilen:")
                tmp = df[["Artikel-Nr.", "Artikelbezeichnung"]].copy()
                tmp.insert(0, "Zeile", range(1, len(tmp) + 1))
                print(tmp.to_string(index=False))

                while True:
                    auswahl = input(
                        f"{Fore.CYAN}Zeilen/Bereiche eingeben "
                        f"(z.B. 1-10, 17-25 | h=Help | Enter=fertig):{Style.RESET_ALL} "
                    ).strip().lower()

                    if auswahl == "":
                        print(f"{Fore.GREEN}Auswahl beendet.{Style.RESET_ALL}")
                        break
                    if auswahl == "h":
                        print(select_help)
                        continue

                    try:
                        rows_1_based = parse_ranges(auswahl)
                        max_row = len(df)
                        rows = [r - 1 for r in rows_1_based if 1 <= r <= max_row]
                        if not rows:
                            print(f"{Fore.RED}Keine g√ºltigen Zeilen ausgew√§hlt.{Style.RESET_ALL}")
                            continue
                    except Exception:
                        print(f"{Fore.RED}Ung√ºltige Eingabe ‚Äì 'h' f√ºr Hilfe.{Style.RESET_ALL}")
                        continue

                    lagerort = input(f"{Fore.YELLOW}Lagerort f√ºr diese Auswahl (leer = keine √Ñnderung):{Style.RESET_ALL} ").strip()
                    sonstiger_text = input(f"{Fore.YELLOW}'Sonstiger Text' f√ºr diese Auswahl (leer = keine √Ñnderung):{Style.RESET_ALL} ").strip()

                    if not lagerort and not sonstiger_text:
                        print(f"{Fore.RED}Keine √Ñnderungen eingegeben ‚Äì Auswahl √ºbersprungen.{Style.RESET_ALL}")
                        continue

                    if lagerort:
                        df.loc[rows, "Lagerort"] = lagerort
                    if sonstiger_text:
                        df.loc[rows, "Sonstiger Text"] = sonstiger_text

            elif answer == "n":
                print(f"{Fore.GREEN}Keine √Ñnderungen an Lagerort / Sonstiger Text.{Style.RESET_ALL}")

        # Artikelnummern immer behandeln
        df["Artikel-Nr."] = df["Artikel-Nr."].astype(str)

        # --- " S" NUR f√ºr MB + Verkaufsware ---
        if kunde == "standard":
            mask_verkauf = df["Sonstiger Text"].str.lower() == "verkaufsware"
            df.loc[mask_verkauf, "Artikel-Nr."] = df.loc[mask_verkauf, "Artikel-Nr."] + " S"

        # --- Prefix je Kunde ---
        if kunde == "ng":
            prefix = "NG "
        elif kunde == "nef":
            prefix = "NEF "
        else:
            prefix = "MB "

        df["Artikel-Nr."] = prefix + df["Artikel-Nr."]
        
        # DEBUG # print("DEBUG Artikel-Nr.:", df["Artikel-Nr."].head().tolist())

    except EOFError:
        prefix = "NG " if kunde == "ng" else ("NEF " if kunde == "nef" else "MB ")
        df["Artikel-Nr."] = prefix + df["Artikel-Nr."].astype(str)

    # Zusatzzeile f√ºr Excel
    try:
        ans_extra = input("M√∂chten Sie eine zus√§tzliche Zeile in der Excel-Datei hinterlegen? (j/n): ").strip().lower()
        if ans_extra == "j":
            extra_excel_line = input("Bitte geben Sie den gew√ºnschten Text f√ºr die Excel-Kopfzeile ein: ").strip()
        else:
            extra_excel_line = ""
    except EOFError:
        extra_excel_line = ""

    # =====================================
    # Index & Spaltenreihenfolge
    # =====================================
    df = df.reset_index(drop=True)
    if "Nr." not in df.columns:
        df.insert(0, "Nr.", df.index + 1)
    else:
        df["Nr."] = range(1, len(df) + 1)

    if kunde == "nef":
        output_columns = [
            "Nr.",
            "Artikel-Nr.",
            "Artikelbezeichnung",
            "LG ID",
            "Charge",
            "Menge PS",
            "Einheit",
            "MHD",
            "Gewicht kg",
            "Lagerort",
        ]
    else:
        output_columns = [
            "Nr.",
            "Artikel-Nr.",
            "Artikelbezeichnung",
            "LG ID",
            "Charge",
            "Menge PS",
            "Einheit",
            "Lademittel",
            "MHD",
            "Gewicht kg",
            "Lagerort",
            "Sonstiger Text",
        ]

    df = df[output_columns]


    # Gewicht in CSV wieder mit Komma
    df["Gewicht kg"] = df["Gewicht kg"].apply(lambda x: str(x).replace(".", ","))

    # LG ID sicher numerisch (verhindert f√ºhrendes "'" in Excel)
    df["LG ID"] = pd.to_numeric(df["LG ID"], errors="coerce").fillna(0).astype(int)

    # =====================================
    # CSV speichern
    # =====================================
    df.to_csv(OUTPUT_CSV, index=False, sep=";", encoding="utf-8-sig")
    print(Fore.GREEN + f"CSV geschrieben nach: {OUTPUT_CSV}")

    # =====================================
    # Excel speichern
    # =====================================
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        sheet_name = "Ruecktour"

        start_row = 2 if extra_excel_line else 0
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        if extra_excel_line:
            ws.cell(row=1, column=1, value=extra_excel_line)

        # Sicherheitsnetz: mindestens ein sichtbares Sheet
        if not wb.worksheets:
            wb.create_sheet(title=sheet_name)
        if all(sh.sheet_state != "visible" for sh in wb.worksheets):
            wb.worksheets[0].sheet_state = "visible"

        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)

        ws.print_area = f"A1:{last_col_letter}{max_row}"

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column_cells), default=10)
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    print(Fore.GREEN + f"Excel-Datei geschrieben nach: {OUTPUT_XLSX}")


if __name__ == "__main__":
    print(Fore.GREEN + START_BANNER)
    print(Fore.YELLOW + f"UnivImport Version {VERSION} mlu")
    ensure_latest_version()
    main()
